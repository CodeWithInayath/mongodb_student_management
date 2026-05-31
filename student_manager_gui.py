import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from pymongo import MongoClient
import csv

# Optional Excel support
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


# ---------- DATABASE CONNECTION ----------
client     = MongoClient("mongodb://localhost:27017/")
db         = client["college"]
collection = db["students"]
collection.create_index([("roll", 1), ("course", 1)], unique=True)


# ---------- THEME ----------
BG       = "#0f1117"
PANEL    = "#1a1d27"
CARD     = "#21253a"
ACCENT   = "#4f8ef7"
ACCENT2  = "#7c5cfc"
SUCCESS  = "#22c97b"
DANGER   = "#ff4f6a"
WARNING  = "#f5a623"
TEXT     = "#e8eaf6"
SUBTEXT  = "#7b80a0"
BORDER   = "#2d3150"
ENTRY_BG = "#171b2e"
ERR_CLR  = "#ff4f6a"
OK_CLR   = "#22c97b"

FONT_LABEL = ("Courier New", 10, "bold")
FONT_ENTRY = ("Courier New", 11)
FONT_BTN   = ("Courier New", 10, "bold")
FONT_TABLE = ("Courier New", 10)
FONT_SUB   = ("Courier New", 9)
FONT_STAT  = ("Georgia", 20, "bold")
FONT_STAT2 = ("Courier New", 9)


# ---------- ROOT ----------
root = tk.Tk()
root.title("Student Management System")
root.geometry("1100x660")
root.configure(bg=BG)
root.resizable(True, True)


# ---------- VARIABLES ----------
roll_var   = tk.StringVar()
name_var   = tk.StringVar()
age_var    = tk.StringVar()
course_var = tk.StringVar()
search_var = tk.StringVar()
filter_var = tk.StringVar(value="All Courses")
status_var = tk.StringVar(value="Ready")


# ---------- STYLE ----------
style = ttk.Style()
style.theme_use("clam")
style.configure("Treeview",
    background=CARD, foreground=TEXT, rowheight=36,
    fieldbackground=CARD, borderwidth=0, font=FONT_TABLE)
style.configure("Treeview.Heading",
    background=PANEL, foreground=ACCENT, relief="flat",
    font=("Courier New", 10, "bold"), padding=8)
style.map("Treeview",
    background=[("selected", ACCENT2)],
    foreground=[("selected", "#ffffff")])
style.layout("Treeview", [("Treeview.treearea", {"sticky": "nswe"})])
style.configure("Vertical.TScrollbar",
    background=PANEL, troughcolor=BG,
    arrowcolor=SUBTEXT, bordercolor=BORDER)
style.configure("TCombobox",
    fieldbackground=ENTRY_BG, background=PANEL,
    foreground=TEXT, arrowcolor=ACCENT,
    selectbackground=ACCENT2, selectforeground=TEXT)
style.map("TCombobox",
    fieldbackground=[("readonly", ENTRY_BG)],
    foreground=[("readonly", TEXT)])


# ---------- HELPERS ----------

def _darken(hex_color):
    r, g, b = int(hex_color[1:3],16), int(hex_color[3:5],16), int(hex_color[5:7],16)
    f = 0.80
    return "#{:02x}{:02x}{:02x}".format(int(r*f), int(g*f), int(b*f))

def make_button(parent, text, color, command, width=13):
    btn = tk.Button(parent, text=text, bg=color, fg="#ffffff",
                    font=FONT_BTN, relief="flat", bd=0,
                    activebackground=color, activeforeground="#ffffff",
                    cursor="hand2", width=width, pady=7, command=command)
    btn.bind("<Enter>", lambda e: btn.config(bg=_darken(color)))
    btn.bind("<Leave>", lambda e: btn.config(bg=color))
    return btn

def set_status(msg, color=SUBTEXT):
    status_var.set(f"  ●  {msg}")
    status_lbl.config(fg=color)


# ---------- INLINE VALIDATION ----------

validation_hints = {}  # key -> tk.Label

def _hint(field, msg, ok=False):
    lbl = validation_hints.get(field)
    if lbl:
        lbl.config(text=msg, fg=OK_CLR if ok else ERR_CLR)

def _clear_hints():
    for lbl in validation_hints.values():
        lbl.config(text="")

def _validate_roll(*_):
    v = roll_var.get().strip()
    if not v: _hint("roll", ""); return
    _hint("roll", "✓" if len(v) >= 2 else "Too short", ok=len(v) >= 2)

def _validate_name(*_):
    v = name_var.get().strip()
    if not v: _hint("name", ""); return
    valid = len(v) >= 2 and all(c.isalpha() or c.isspace() for c in v)
    if len(v) < 2:        _hint("name", "Too short")
    elif not valid:        _hint("name", "Letters only")
    else:                  _hint("name", "✓", ok=True)

def _validate_age(*_):
    v = age_var.get().strip()
    if not v: _hint("age", ""); return
    try:
        age = int(v)
        if 10 <= age <= 80: _hint("age", "✓", ok=True)
        else:               _hint("age", "Must be 10–80")
    except ValueError:      _hint("age", "Numbers only")

def _validate_course(*_):
    v = course_var.get().strip()
    if not v: _hint("course", ""); return
    _hint("course", "✓", ok=True)

roll_var.trace_add("write",   _validate_roll)
name_var.trace_add("write",   _validate_name)
age_var.trace_add("write",    _validate_age)
course_var.trace_add("write", _validate_course)


# ---------- CORE FUNCTIONS ----------

def clear_fields():
    roll_var.set(""); name_var.set("")
    age_var.set("");  course_var.set("")
    _clear_hints()
    set_status("Fields cleared")

def get_all_courses():
    return ["All Courses"] + sorted(collection.distinct("course"))

def refresh_course_filter():
    current = filter_var.get()
    courses = get_all_courses()
    course_filter["values"] = courses
    if current not in courses:
        filter_var.set("All Courses")

def build_query():
    q    = {}
    text = search_var.get().strip()
    crs  = filter_var.get()
    if text:
        rx = {"$regex": text, "$options": "i"}
        q["$or"] = [{"roll": rx}, {"name": rx}, {"course": rx}]
    if crs and crs != "All Courses":
        q["course"] = crs
    return q

def refresh_table():
    student_table.delete(*student_table.get_children())
    count = 0
    for s in collection.find(build_query()):
        student_table.insert("", "end", values=(
            s.get("roll","—"), s.get("name","—"),
            s.get("age","—"),  s.get("course","—")
        ))
        count += 1
    total_lbl.config(text=f"  {count} record{'s' if count != 1 else ''} found")
    update_stats()

def update_stats():
    total   = collection.count_documents({})
    courses = len(collection.distinct("course"))
    ages    = [a["age"] for a in collection.find({}, {"age":1,"_id":0})
               if isinstance(a.get("age"), int)]
    avg_age = round(sum(ages)/len(ages), 1) if ages else "—"

    stat_total_val.config(text=str(total))
    stat_courses_val.config(text=str(courses))
    stat_avg_val.config(text=str(avg_age))

    result = list(collection.aggregate([
        {"$group": {"_id": "$course", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}}, {"$limit": 1}
    ]))
    top = result[0]["_id"] if result else "—"
    stat_top_val.config(text=top[:14] + ("…" if len(top) > 14 else ""))

def _validate_form():
    if not all([roll_var.get().strip(), name_var.get().strip(),
                age_var.get().strip(), course_var.get().strip()]):
        return None, "All fields are required."
    if not all(c.isalpha() or c.isspace() for c in name_var.get().strip()):
        return None, "Name must contain letters only."
    try:
        age = int(age_var.get().strip())
        if not (10 <= age <= 80): raise ValueError
    except ValueError:
        return None, "Age must be a number between 10 and 80."
    return age, None

def add_student():
    age, err = _validate_form()
    if err: messagebox.showerror("Validation Error", err); return
    if collection.find_one({"roll": roll_var.get().strip(), "course": course_var.get().strip()}):
        messagebox.showerror("Duplicate", f"Roll No '{roll_var.get()}' already exists in {course_var.get()}."); return
    collection.insert_one({
        "roll": roll_var.get().strip(), "name": name_var.get().strip(),
        "age": age, "course": course_var.get().strip()
    })
    refresh_table(); refresh_course_filter(); clear_fields()
    set_status("Student added successfully.", SUCCESS)

def update_student():
    selected = student_table.focus()
    if not selected: messagebox.showwarning("No Selection", "Select a student first."); return
    age, err = _validate_form()
    if err: messagebox.showerror("Validation Error", err); return
    if not messagebox.askyesno("Confirm Update", f"Update '{name_var.get()}'?"): return
    old_roll   = student_table.item(selected, "values")[0]
    old_course = student_table.item(selected, "values")[3]
    # Check if the new roll+course combo already exists on a *different* record
    conflict = collection.find_one({
        "roll":   roll_var.get().strip(),
        "course": course_var.get().strip(),
        "$nor":   [{"roll": old_roll, "course": old_course}]
    })
    if conflict:
        messagebox.showerror("Duplicate",
            f"Roll No '{roll_var.get()}' already exists in {course_var.get()}."); return
    collection.update_one({"roll": old_roll, "course": old_course}, {"$set": {
        "roll": roll_var.get().strip(), "name": name_var.get().strip(),
        "age": age, "course": course_var.get().strip()
    }})
    refresh_table(); refresh_course_filter()
    set_status("Student record updated.", WARNING)

def delete_student():
    selected = student_table.focus()
    if not selected: messagebox.showwarning("No Selection", "Select a student first."); return
    name = student_table.item(selected, "values")[1]
    if messagebox.askyesno("Confirm Delete", f"Delete '{name}'? This cannot be undone."):
        roll = student_table.item(selected, "values")[0]
        collection.delete_one({"roll": roll})
        refresh_table(); refresh_course_filter(); clear_fields()
        set_status(f"'{name}' deleted.", DANGER)

def select_student(event):
    selected = student_table.focus()
    values   = student_table.item(selected, "values")
    if values:
        roll_var.set(values[0]); name_var.set(values[1])
        age_var.set(values[2]);  course_var.set(values[3])

def do_search(*_): refresh_table()
def do_filter(*_): refresh_table()


# ---------- EXPORT ----------

def export_csv():
    data = list(collection.find(build_query()))
    if not data: messagebox.showinfo("No Data", "Nothing to export."); return
    path = filedialog.asksaveasfilename(
        defaultextension=".csv", filetypes=[("CSV files","*.csv")], title="Save as CSV")
    if not path: return
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["Roll No","Name","Age","Course"])
        for s in data:
            w.writerow([s.get("roll",""), s.get("name",""),
                        s.get("age",""), s.get("course","")])
    set_status(f"Exported {len(data)} rows to CSV.", SUCCESS)
    messagebox.showinfo("Export Successful", f"Saved to:\n{path}")

def export_excel():
    if not EXCEL_AVAILABLE:
        messagebox.showerror("Missing Library",
            "openpyxl is not installed.\nRun:  pip install openpyxl"); return
    data = list(collection.find(build_query()))
    if not data: messagebox.showinfo("No Data", "Nothing to export."); return
    path = filedialog.asksaveasfilename(
        defaultextension=".xlsx", filetypes=[("Excel files","*.xlsx")], title="Save as Excel")
    if not path: return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Students"
    hfill = PatternFill("solid", fgColor="1a1d27")
    hfont = Font(bold=True, color="4f8ef7", name="Courier New")
    for col, h in enumerate(["Roll No","Name","Age","Course"], 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hfill; c.font = hfont
        c.alignment = Alignment(horizontal="center")
    for i, s in enumerate(data, 2):
        for col, val in enumerate([s.get("roll",""), s.get("name",""),
                                    s.get("age",""), s.get("course","")], 1):
            c = ws.cell(row=i, column=col, value=val)
            c.fill = PatternFill("solid", fgColor="21253a" if i%2==0 else "171b2e")
            c.alignment = Alignment(horizontal="center")
    for col, w in zip("ABCD", [14,22,8,20]):
        ws.column_dimensions[col].width = w
    wb.save(path)
    set_status(f"Exported {len(data)} rows to Excel.", SUCCESS)
    messagebox.showinfo("Export Successful", f"Saved to:\n{path}")


# ═══════════════════════════════════════════════════════════════
# LAYOUT
# ═══════════════════════════════════════════════════════════════

# ── LEFT SIDEBAR ────────────────────────────────────────────────
sidebar = tk.Frame(root, bg=PANEL, width=300)
sidebar.pack(side="left", fill="y")
sidebar.pack_propagate(False)

hdr = tk.Frame(sidebar, bg=PANEL, pady=18)
hdr.pack(fill="x", padx=20)
tk.Label(hdr, text="SMS", font=("Georgia", 28, "bold"), bg=PANEL, fg=ACCENT).pack(anchor="w")
tk.Label(hdr, text="Student Management System", font=FONT_SUB, bg=PANEL, fg=SUBTEXT).pack(anchor="w")
tk.Frame(sidebar, bg=BORDER, height=1).pack(fill="x", padx=20)

# ── FORM ────────────────────────────────────────────────────────
form = tk.Frame(sidebar, bg=PANEL, padx=20)
form.pack(fill="x", pady=8)

def add_field(parent, label, var, key):
    row = tk.Frame(parent, bg=PANEL)
    row.pack(fill="x", pady=(6, 0))
    top = tk.Frame(row, bg=PANEL)
    top.pack(fill="x")
    tk.Label(top, text=label, bg=PANEL, fg=SUBTEXT, font=FONT_LABEL).pack(side="left")
    hint_lbl = tk.Label(top, text="", bg=PANEL, fg=ERR_CLR, font=FONT_SUB)
    hint_lbl.pack(side="right")
    validation_hints[key] = hint_lbl
    border = tk.Frame(row, bg=BORDER)
    border.pack(fill="x", pady=(2,0))
    entry = tk.Entry(border, textvariable=var, bg=ENTRY_BG, fg=TEXT,
                     font=FONT_ENTRY, insertbackground=ACCENT,
                     relief="flat", bd=7, highlightthickness=0)
    entry.pack(fill="x", padx=1, pady=1)
    entry.bind("<FocusIn>",  lambda e: border.config(bg=ACCENT))
    entry.bind("<FocusOut>", lambda e: border.config(bg=BORDER))

add_field(form, "ROLL NUMBER", roll_var,   "roll")
add_field(form, "FULL NAME",   name_var,   "name")
add_field(form, "AGE",         age_var,    "age")
add_field(form, "COURSE",      course_var, "course")

# ── ACTION BUTTONS ──────────────────────────────────────────────
btn_area = tk.Frame(sidebar, bg=PANEL, padx=20)
btn_area.pack(fill="x", pady=10)

r1 = tk.Frame(btn_area, bg=PANEL)
r1.pack(fill="x")
make_button(r1, "＋  ADD",    SUCCESS, add_student).pack(side="left", padx=(0,6))
make_button(r1, "✎  UPDATE", WARNING, update_student).pack(side="left")

r2 = tk.Frame(btn_area, bg=PANEL)
r2.pack(fill="x", pady=(6,0))
make_button(r2, "✕  DELETE", DANGER,  delete_student).pack(side="left", padx=(0,6))
make_button(r2, "⟳  CLEAR",  SUBTEXT, clear_fields).pack(side="left")

tk.Frame(sidebar, bg=BORDER, height=1).pack(fill="x", padx=20, pady=10)

# ── EXPORT BUTTONS ──────────────────────────────────────────────
tk.Label(sidebar, text="EXPORT CURRENT VIEW", bg=PANEL,
         fg=SUBTEXT, font=FONT_LABEL, anchor="w").pack(fill="x", padx=20, pady=(0,6))

exp_row = tk.Frame(sidebar, bg=PANEL, padx=20)
exp_row.pack(fill="x")
make_button(exp_row, "⬇  CSV",   ACCENT,  export_csv).pack(side="left", padx=(0,6))
make_button(exp_row, "⬇  EXCEL", ACCENT2, export_excel).pack(side="left")

tk.Frame(sidebar, bg=BORDER, height=1).pack(fill="x", padx=20, pady=10)

status_lbl = tk.Label(sidebar, textvariable=status_var, font=FONT_SUB,
                       bg=PANEL, fg=SUBTEXT, anchor="w")
status_lbl.pack(fill="x", padx=20)


# ── RIGHT PANEL ─────────────────────────────────────────────────
right = tk.Frame(root, bg=BG)
right.pack(side="right", fill="both", expand=True)

# ── STATS CARDS ─────────────────────────────────────────────────
stats_bar = tk.Frame(right, bg=BG)
stats_bar.pack(fill="x", padx=24, pady=(16, 0))

def stat_card(parent, label, default="0"):
    card = tk.Frame(parent, bg=CARD, padx=18, pady=12)
    card.pack(side="left", padx=(0,10), fill="y")
    val = tk.Label(card, text=default, font=FONT_STAT, bg=CARD, fg=TEXT)
    val.pack(anchor="w")
    tk.Label(card, text=label, font=FONT_STAT2, bg=CARD, fg=SUBTEXT).pack(anchor="w")
    return val

stat_total_val   = stat_card(stats_bar, "Total Students")
stat_courses_val = stat_card(stats_bar, "Unique Courses")
stat_avg_val     = stat_card(stats_bar, "Average Age", "—")
stat_top_val     = stat_card(stats_bar, "Top Course",  "—")

# ── TOPBAR ──────────────────────────────────────────────────────
topbar = tk.Frame(right, bg=BG, pady=12)
topbar.pack(fill="x", padx=24)

tk.Label(topbar, text="All Students", font=("Georgia", 14, "bold"),
         bg=BG, fg=TEXT).pack(side="left")
total_lbl = tk.Label(topbar, text="", font=FONT_SUB, bg=BG, fg=SUBTEXT)
total_lbl.pack(side="left", padx=10)

# Course filter
filter_frame = tk.Frame(topbar, bg=BORDER, padx=1, pady=1)
filter_frame.pack(side="right", padx=(6,0))
course_filter = ttk.Combobox(filter_frame, textvariable=filter_var,
                              state="readonly", width=18, font=FONT_ENTRY)
course_filter["values"] = get_all_courses()
course_filter.pack(padx=1, pady=1)
course_filter.bind("<<ComboboxSelected>>", do_filter)
tk.Label(topbar, text="Course:", font=FONT_LABEL, bg=BG, fg=SUBTEXT).pack(side="right", padx=(0,4))

# Search
search_frame = tk.Frame(topbar, bg=BORDER, padx=1, pady=1)
search_frame.pack(side="right", padx=(0,12))
tk.Label(search_frame, text="⌕", font=("Courier New", 13),
         bg=ENTRY_BG, fg=SUBTEXT).pack(side="left", padx=(8,0))
tk.Entry(search_frame, textvariable=search_var, width=20, bg=ENTRY_BG,
         fg=TEXT, insertbackground=ACCENT, font=FONT_ENTRY,
         relief="flat", bd=6).pack(side="left")
search_var.trace_add("write", do_search)

# ── TABLE ────────────────────────────────────────────────────────
table_wrap = tk.Frame(right, bg=BG, padx=24)
table_wrap.pack(fill="both", expand=True, pady=(0,6))

cols = ("Roll No", "Name", "Age", "Course")
student_table = ttk.Treeview(table_wrap, columns=cols, show="headings", selectmode="browse")
for col in cols:
    student_table.heading(col, text=col.upper())
    student_table.column(col, anchor="center", minwidth=60,
                         width=80 if col == "Age" else 170)

vsb = ttk.Scrollbar(table_wrap, orient="vertical", command=student_table.yview)
student_table.configure(yscrollcommand=vsb.set)
student_table.pack(side="left", fill="both", expand=True)
vsb.pack(side="right", fill="y")
student_table.bind("<ButtonRelease-1>", select_student)

# Footer
footer = tk.Frame(right, bg=BG, pady=5)
footer.pack(fill="x", padx=24)
tk.Label(footer,
         text="Click a row to select  ·  Search by roll / name / course  ·  Filter by course  ·  Export exports current filtered view",
         font=FONT_SUB, bg=BG, fg=SUBTEXT).pack(side="left")


# ---------- INIT ----------
refresh_table()
refresh_course_filter()
set_status("Connected to MongoDB  ●  college.students")
root.mainloop()